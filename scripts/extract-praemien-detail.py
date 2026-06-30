#!/usr/bin/env python3
"""Extract detailed BAG premiums (per insurer, canton, region, franchise) from gesamtbericht_ch.xlsx.

Erfasst beide Unfall-Tarife des Standardmodells (TAR-BASE):
  MIT-UNF (mit Unfalldeckung) → e/j/k  (Erwachsene/Junge/Kinder)
  OHN-UNF (ohne Unfalldeckung) → eo/jo (nur Erwachsene/Junge; Kinder wählen nicht
            über einen Arbeitgeber ab, daher kein OHN-UNF für Kinder)
"""
import openpyxl
import json
import sys
import os

raw_dir = sys.argv[1] if len(sys.argv) > 1 else os.path.join(os.path.dirname(__file__), '..', 'src', 'data', 'raw')

wb = openpyxl.load_workbook(os.path.join(raw_dir, 'gesamtbericht_ch.xlsx'), read_only=True, data_only=True)
ws = wb['Export']

# Insurer name mapping from zugelassene_kv
wb2 = openpyxl.load_workbook(os.path.join(raw_dir, 'zugelassene_kv.xlsx'), data_only=True)
ws2 = wb2[wb2.sheetnames[1]]
nr_to_name = {}
for row in ws2.iter_rows(min_row=4, values_only=False):
    nr = row[1].value
    x = row[2].value
    name = row[3].value
    if nr and name and not x:
        nr_to_name[int(nr)] = str(name).strip()

# ERW/JUG franchise order (fixed): 300, 500, 1000, 1500, 2000, 2500
ERW_FRA = ['FRA-300', 'FRA-500', 'FRA-1000', 'FRA-1500', 'FRA-2000', 'FRA-2500']
# KIN franchise order (fixed): 0, 100, 200, 300, 400, 500, 600
KIN_FRA = ['FRA-0', 'FRA-100', 'FRA-200', 'FRA-300', 'FRA-400', 'FRA-500', 'FRA-600']

era_idx = {f: i for i, f in enumerate(ERW_FRA)}
kin_idx = {f: i for i, f in enumerate(KIN_FRA)}

# Collect: key = (versNr, kanton, regionNr) → {e,j,k (mit) + eo,jo (ohne)}
data = {}

region_map = {'PR-REG CH0': 0, 'PR-REG CH1': 1, 'PR-REG CH2': 2, 'PR-REG CH3': 3}

for row in ws.iter_rows(min_row=2, values_only=True):
    tarif_type = row[9]
    unfall = row[7]
    if tarif_type != 'TAR-BASE' or unfall not in ('MIT-UNF', 'OHN-UNF'):
        continue

    age_class = row[6]
    age_sub = row[10] or ''
    if age_class == 'AKL-KIN' and age_sub != 'K1':
        continue

    vers_nr = int(row[0])
    kanton = row[1]
    region = region_map.get(row[5], 0)
    franchise = row[12]
    premium = round(float(row[13]), 1) if row[13] else 0

    key = (vers_nr, kanton, region)
    if key not in data:
        data[key] = {'e': [0]*6, 'j': [0]*6, 'k': [0]*7, 'eo': [0]*6, 'jo': [0]*6}

    entry = data[key]
    mit = (unfall == 'MIT-UNF')
    if age_class == 'AKL-ERW' and franchise in era_idx:
        entry['e' if mit else 'eo'][era_idx[franchise]] = premium
    elif age_class == 'AKL-JUG' and franchise in era_idx:
        entry['j' if mit else 'jo'][era_idx[franchise]] = premium
    elif age_class == 'AKL-KIN' and franchise in kin_idx and mit:
        entry['k'][kin_idx[franchise]] = premium

# Build output: nested by versNr → array of [kanton, region, e[], j[], k[], eo[], jo[]]
output = {}
for (vers_nr, kanton, region), vals in sorted(data.items()):
    vk = str(vers_nr)
    if vk not in output:
        output[vk] = []
    output[vk].append([kanton, region, vals['e'], vals['j'], vals['k'], vals['eo'], vals['jo']])

print(json.dumps({
    'insurers': nr_to_name,
    'premiums': output,
    'erwFranchises': [300, 500, 1000, 1500, 2000, 2500],
    'kinFranchises': [0, 100, 200, 300, 400, 500, 600]
}))
