#!/usr/bin/env python3
"""Extract BAG Prämienregionen + Versichererliste from XLSX to JSON."""
import openpyxl
import json
import sys
import os

raw_dir = sys.argv[1] if len(sys.argv) > 1 else os.path.join(os.path.dirname(__file__), '..', 'src', 'data', 'raw')

wb = openpyxl.load_workbook(os.path.join(raw_dir, 'praemienregionen.xlsx'), data_only=True)
ws = wb['D_PRIM']

praemien = {}
for row in ws.iter_rows(min_row=17, values_only=True):
    bfs, kanton, gemeinde, region, kinder, junge, erwachsene = row[:7]
    if bfs and kanton:
        bfs_str = str(int(bfs) if isinstance(bfs, float) else bfs)
        praemien[bfs_str] = {
            "k": str(kanton),
            "g": str(gemeinde),
            "r": int(region) if region else 0,
            "c": round(float(kinder), 1) if kinder else 0,
            "y": round(float(junge), 1) if junge else 0,
            "a": round(float(erwachsene), 1) if erwachsene else 0
        }

wb2 = openpyxl.load_workbook(os.path.join(raw_dir, 'zugelassene_kv.xlsx'), data_only=True)
ws2 = wb2[wb2.sheetnames[1]]

insurers = []
for row in ws2.iter_rows(min_row=4, values_only=False):
    nr = row[1].value
    x = row[2].value
    name = row[3].value
    ort = row[4].value
    if nr and name and not x:
        insurers.append({"nr": int(nr), "name": str(name).strip(), "ort": str(ort).strip() if ort else ""})

insurers.sort(key=lambda x: x['name'].lower())

print(json.dumps({"praemien": praemien, "insurers": insurers}))
